#!/usr/bin/env python
# encoding: utf-8
import shutil
import os.path
import zipfile
import logging
import time
import xlrd
from openpyxl import load_workbook
import datetime
from PIL import Image, ImageFont, ImageDraw
import chardet


# TODO Add the extra function
def generate_logging():
    """
    return:
        The logger output the log message

    """
    # First, generate a logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Second, generate a log handler to write the log file
    rq = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
    log_path = os.getcwd() + '/Logs/'
    if os.path.exists(log_path):
        pass
    else:
        os.mkdir(log_path)

    log_name = log_path + rq + '.log'
    logfile = log_name
    fh = logging.FileHandler(logfile, mode='w')
    fh.setLevel(logging.DEBUG)

    # Third, define the output format of handler
    formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
    fh.setFormatter(formatter)

    # Four, add the handler into logger
    logger.addHandler(fh)

    return logger


class ManageClass:
    def __init__(self, zip_file_path, excel_file_path):
        self.zip_file_path = zip_file_path
        self.excel_file_path = excel_file_path
        self.logger = generate_logging()
        self.excel_value = []
        self.dec_path = ''
        self.target_path = ''

    def get_zip_file(self, dec_path='./tmp/'):
        try:
            self.dec_path = dec_path
            if os.path.exists(self.dec_path):
                shutil.rmtree(self.dec_path)
            f = zipfile.ZipFile(self.zip_file_path, "r")

            if os.path.exists(dec_path):
                pass
            else:
                os.mkdir(dec_path)

            for file in f.namelist():
                f.extract(file, dec_path)

            self.target_path = os.path.join(os.path.dirname(self.zip_file_path),
                                      datetime.datetime.now().strftime("%Y%m%d"),
                                      os.path.splitext(os.path.basename(self.zip_file_path))[0])
            if os.path.exists(self.target_path):
                pass
            else:
                os.makedirs(self.target_path)

            self.logger.info('解压压缩文件成功')
            return [True, '']
        except Exception as e:
            self.logger.warning('解压压缩文件失败，失败原因：{0}'.format(e))
            return [False, e]

    def get_excel_file(self):
        try:
            extension_name = os.path.splitext(self.excel_file_path)[1]
            if extension_name == '.xls':
                workbook = xlrd.open_workbook(self.excel_file_path)
                booksheet = workbook.sheet_by_index(0)

                rowNum = booksheet.nrows

                for index_row in range(2, rowNum + 1):
                    value = []
                    for index_col in range(1, 5):
                        value.append(booksheet.cell_value(index_row, index_col))
                    self.excel_value.append(value)
                self.logger.info('读取Excel文件内容成功!')
                return True
            else:
                workbook = load_workbook(self.excel_file_path)
                # sheets = workbook.get_sheet_names()
                booksheet = workbook.worksheets[0]

                rowNum = booksheet.max_row

                for index_row in range(2, rowNum + 1):
                    value = []
                    for index_col in range(1, 5):
                        value.append(booksheet.cell(row=index_row, column=index_col).value)
                    self.excel_value.append(value)
                self.logger.info('读取Excel文件内容成功!')

                return [True, '']
        except Exception as e:
            self.excel_value = []
            self.logger.warning('读取Excel文件内容失败，失败原因：{0}'.format(e))
            return [False, e]

    def coding_conversion(self):
        for index, file_name in enumerate(os.listdir(self.dec_path)):
            try:
                os.rename(os.path.join(self.dec_path, file_name)
                          , os.path.join(self.dec_path, file_name.encode('cp437').decode('gbk')))
            except Exception as e:
                print(e)
            # if index == 0:
            #     os.rename(os.path.join(self.dec_path, file_name)
            #               , os.path.join(self.dec_path, '何友娣南街菜场_12_6230910299054016641_20201222193838.png'))
            # else:
            #     os.rename(os.path.join(self.dec_path, file_name)
            #               , os.path.join(self.dec_path, '宁波市镇海区庄市月光族服装店_12_6230910299043021074_20201222193838.png'))
            # os.rename(os.path.join(self.dec_path, file_name)
            #           , os.path.join(self.dec_path, random.randint(0, 10000) + '.jpg'))

    def convert_image(self, filename):
        try:
            # 定义背景宽度像素
            width = 1183
            # 定义背景高度像素
            # high = 1676
            # 定义粘贴后的二维码大小

            # 老
            # ewmwidth = 680
            # ewmhigh = 680
            # 新
            ewmwidth = 960
            ewmhigh = 960
            # 二维码粘贴高度定位为580，小牌下移0；大牌下移221

            # 老
            # gd = 580
            # tz = 221
            # 新
            gd = 405
            tz = 221

            # 定义字体大小
            # 老
            # zt = 88
            # 新
            zt = 80

            # 打开底版图片
            im = Image.open('./template.jpg')
            # 打开二维码图片
            imin = Image.open(self.dec_path + filename)
            # 缩放
            imin = imin.resize((ewmwidth, ewmhigh))
            # 粘贴图片
            # 老
            # im.paste(imin, (int((width - ewmwidth) / 2), gd + tz))
            # 新
            im.paste(imin, (int((width - ewmwidth) / 2), gd))

            name = ''
            save_file_name = ''
            if self.excel_value:
                card_number = filename.split('_')[2]
                for item in self.excel_value:
                    if card_number == item[3]:
                        name = item[1] if item[1] else item[0]
                        save_file_name = item[0]
                        # name = item[0]
                        break

                if not name:
                    name = filename.split('_')[0]
                    save_file_name = filename.split('_')[0]
                else:
                    pass
            else:
                name = filename.split('_')[0]
                save_file_name = filename.split('_')[0]
            length = len(name)

            # 使用自定义的字体，第二个参数表示字符大小
            font = ImageFont.truetype('./HeiTi.TTF', zt)
            # 在图片上添加文字
            draw = ImageDraw.Draw(im)

            # 限制输入16个字符
            # 限制每行字数为8，即8个字换行
            # 老
            # hangzishu = 8
            # 新
            hangzishu = 10
            if length <= hangzishu:  # 字符串很长的情况下
                an = (width - length * zt) / 2  # 判断字符串到图片左侧的距离
                # 老
                # draw.text((an, 1285 + tz), name, fill=(0, 0, 0), font=font)  # 文字写入
                # 新
                draw.text((an, 1457), name, fill=(0, 0, 0), font=font)  # 文字写入
            elif (length > hangzishu) and (length <= hangzishu * 2):
                an1 = (width - hangzishu * zt) / 2  # 第一行
                an2 = (width - (length - hangzishu) * zt) / 2  # 第二行
                a1, a2 = name[:hangzishu], name[hangzishu:]
                # 老
                # draw.text((an1, 1252 + tz), a1, fill=(0, 0, 0), font=font)
                # draw.text((an2, 1349 + tz), a2, fill=(0, 0, 0), font=font)
                # 新
                draw.text((an1, 1410), a1, fill=(0, 0, 0), font=font)
                draw.text((an2, 1507), a2, fill=(0, 0, 0), font=font)
            else:
                # 人为控制字符长度
                self.logger.info('{0}字符超限'.format(filename))

            im.save(os.path.join(self.target_path, save_file_name) + ".jpg", 'JPEG')
            return [True, save_file_name]
        except Exception as e:
            self.logger.warning('二维码转换出现错误,文件名:{0}, {1}'.format(os.path.basename(filename), e))
            return [False, e]

    def convert_image_colours(self, filename):
        try:
            # 定义背景宽度像素
            width = 1183
            # 定义背景高度像素
            # high = 1676
            # 定义粘贴后的二维码大小
            # 老
            # ewmwidth = 580
            # ewmhigh = 580

            # 新
            ewmwidth = 685
            ewmhigh = 670
            # 二维码粘贴高度定位为580，小牌下移0；大牌下移221
            # gd = 580

            # 定义字体大小
            # 老
            # zt = 85

            # 新
            zt = 55

            # 打开底版图片
            im = Image.open('./template_colours_new.jpg')
            # 打开二维码图片
            imin = Image.open(self.dec_path + filename)
            # 缩放
            imin = imin.resize((ewmwidth, ewmhigh))
            # 粘贴图片
            # 老
            # im.paste(imin, (310, 780))
            # 新
            im.paste(imin, (251, 495))

            imdecorate = Image.open('./logo.png')
            r, g, b, a = imdecorate.split()
            png_info = imdecorate.info
            im.paste(imdecorate, (515, 756), mask=a)

            name = ''
            save_file_name = ''
            if self.excel_value:
                card_number = filename.split('_')[2]
                for item in self.excel_value:
                    if card_number == item[3]:
                        name = item[1] if item[1] else item[0]
                        save_file_name = item[0]
                        # name = item[0]
                        break

                if not name:
                    name = filename.split('_')[0]
                    save_file_name = filename.split('_')[0]
                else:
                    pass
            else:
                name = filename.split('_')[0]
                save_file_name = filename.split('_')[0]
            length = len(name)

            # 使用自定义的字体，第二个参数表示字符大小
            font = ImageFont.truetype('./HeiTi.TTF', zt)
            # 在图片上添加文字
            draw = ImageDraw.Draw(im)

            # 限制输入16个字符
            # 限制每行字数为8，即8个字换行
            # 老
            # hangzishu = 8
            # if length <= hangzishu:  # 字符串很长的情况下
            #     an = (width - length * zt) / 2  # 判断字符串到图片左侧的距离
            #     draw.text((an, 1426), name, fill=(0, 0, 0), font=font)  # 文字写入
            # elif (length > hangzishu) and (length <= hangzishu * 2):
            #     an1 = (width - hangzishu * zt) / 2  # 第一行
            #     an2 = (width - (length - hangzishu) * zt) / 2  # 第二行
            #     a1, a2 = name[:hangzishu], name[hangzishu:]
            #     draw.text((an1, 1434), a1, fill=(0, 0, 0), font=font)
            #     draw.text((an2, 1540), a2, fill=(0, 0, 0), font=font)
            # else:
            #     # 人为控制字符长度
            #     self.logger.info('{0}字符超限'.format(filename))

            # 新
            hangzishu = 10
            if length <= hangzishu:  # 字符串很长的情况下
                an = (width - length * zt) / 2  # 判断字符串到图片左侧的距离
                draw.text((an, 1178), name, fill=(0, 0, 0), font=font)  # 文字写入
            elif (length > hangzishu) and (length <= hangzishu * 2):
                an1 = (width - hangzishu * zt) / 2  # 第一行
                an2 = (width - (length - hangzishu) * zt) / 2  # 第二行
                a1, a2 = name[:hangzishu], name[hangzishu:]
                draw.text((an1, 1178), a1, fill=(0, 0, 0), font=font)
                draw.text((an2, 1242), a2, fill=(0, 0, 0), font=font)
            else:
                # 人为控制字符长度
                self.logger.info('{0}字符超限'.format(filename))

            im.save(os.path.join(self.target_path, save_file_name + "_彩色图") + ".jpg", 'JPEG')
            # im.save(os.path.join(self.target_path, save_file_name + "_彩色图") + ".png")
            return [True, save_file_name + "_彩色图"]
        except Exception as e:
            self.logger.warning('二维码转换出现错误,文件名:{0}, {1}'.format(os.path.basename(filename), e))
            return [False, e]

    def return_list_filename(self):
        return os.listdir(self.dec_path)

    def del_temp_file(self):
        if os.path.exists(self.dec_path):
            shutil.rmtree(self.dec_path)
            self.logger.info('删除临时文件夹{0}成功'.format(self.dec_path))

    def return_excel_value(self):
        return self.excel_value


if __name__ == '__main__':
    manage = ManageClass('C:\\Users\\Administrator\\Desktop\\码牌名称隐藏姓名-打印版\\测试.zip',
                         'C:\\Users\\Administrator\\Desktop\\码牌名称隐藏姓名-打印版\\测试-丰收一码通商户信息登记表.xlsx')
    # manage.get_zip_file()
    # list_filename = manage.return_list_filename()
    #
    # for filename in list_filename:
    #     manage.convert_image(filename)
    # 'C:\\Users\\Administrator\\Desktop\\ImageSuperpose\\OneQRCodePassImage_20190102164744.zip'
    manage.get_zip_file()
    manage.get_excel_file()
    manage.coding_conversion()
    print(manage.return_excel_value())
    list_filename = manage.return_list_filename()

    for filename in list_filename:
        manage.convert_image_colours(filename)

    manage.del_temp_file()
